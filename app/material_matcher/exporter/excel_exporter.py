"""Excel 导出器 - 按工料机汇总-安装.xlsx模板格式"""
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from ..models.material import AggregatedMaterial, PriceQuote
from ..utils.logging import get_logger


logger = get_logger(__name__)


class ExcelExporter:
    """Excel 导出器 - 按工料机汇总-安装.xlsx模板格式"""

    # 模板列顺序（15列，按工料机汇总-安装.xlsx）
    TEMPLATE_COLUMNS = [
        '序号',           # A
        '材料名称',       # B
        '规格、型号',      # C
        '单位',           # D
        '数量',           # E
        '预算材料单价',    # F
        '聚鑫预算',       # G
        '概算',           # H
        '施工单位',       # I  (合并单元格I1:K1，三列都是"施工单位")
        '施工单位',       # J
        '施工单位',       # K
        '过控',           # L
        '信息价',         # M
        '预算与概算差价', # N
        '备注',           # O
    ]

    def __init__(self, template_path: Optional[str] = None):
        """
        Args:
            template_path: 模板文件路径（可选，用于复制格式）
        """
        self.template_path = template_path

    def export(
        self,
        materials: list[AggregatedMaterial],
        output_path: str
    ):
        """
        导出到 Excel

        Args:
            materials: 整合后的材料列表
            output_path: 输出文件路径
        """
        logger.info("exporting", count=len(materials), output=output_path)

        # 构建 DataFrame（按模板列顺序）
        rows = []
        for m in materials:
            # 按列顺序构建列表（避免字典键重复问题）
            row = [
                m.id,                                              # 序号
                m.name,                                            # 材料名称
                m.specification or '',                             # 规格、型号
                m.unit or '',                                      # 单位
                m.quantity if m.quantity else '',                  # 数量
                m.budget_price if m.budget_price else '',         # 预算材料单价
                m.budget_price if m.budget_price else '',         # 聚鑫预算
                '',                                                # 概算
                '', '', '',                                        # 施工单位1/2/3（价格）
                m.control_price if m.control_price else '',        # 过控
                '/',                                               # 信息价
                m.price_diff if m.price_diff is not None else '',  # 预算与概算差价
                '',                                                # 备注
            ]

            # 填充施工单位报价（最多3个）- 模板里是价格数值
            for i, q in enumerate(m.quotes[:3]):
                row[8 + i] = q.price  # 施工单位1/2/3 列

            # 备注里放供应商名称
            if m.quotes:
                suppliers = [q.supplier for q in m.quotes[:3]]
                row[14] = '; '.join(suppliers)  # 备注列

            rows.append(row)

        df = pd.DataFrame(rows, columns=self.TEMPLATE_COLUMNS)

        # 保存到 Excel
        output_path = Path(output_path)
        df.to_excel(output_path, index=False, sheet_name='工料机汇总调价表', engine='openpyxl')

        # 设置列宽
        self._set_column_width(output_path)

        logger.info("export_completed", output=str(output_path))

    def _set_column_width(self, file_path: Path):
        """设置列宽和合并单元格"""
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path)
            ws = wb.active

            # 列宽映射（按模板）
            widths = {
                'A': 6,    # 序号
                'B': 28,   # 材料名称
                'C': 16,   # 规格、型号
                'D': 6,    # 单位
                'E': 10,   # 数量
                'F': 12,   # 预算材料单价
                'G': 12,   # 聚鑫预算
                'H': 10,   # 概算
                'I': 30,   # 施工单位
                'J': 30,   # 施工单位2
                'K': 30,   # 施工单位3
                'L': 12,   # 过控
                'M': 10,   # 信息价
                'N': 14,   # 预算与概算差价
                'O': 30,   # 备注
            }

            for col, width in widths.items():
                if col in ws.column_dimensions:
                    ws.column_dimensions[col].width = width

            # 添加合并单元格（表头行）
            # I1:K1 = 施工单位
            ws.merge_cells(start_row=1, start_column=9, end_row=1, end_column=11)

            wb.save(file_path)
        except Exception as e:
            logger.warning("set_column_width_failed", error=str(e))
